#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import shutil
import glob
import subprocess
import sys
from time import time
from openpyxl import load_workbook

wb = load_workbook("params.xlsx", data_only=True)
ws = wb.active       # 默认取第一个 sheet
params = {}

for row in ws.iter_rows(min_row=1, values_only=True):
    key, value = row[0], row[1]
    if key is None:
        continue       # 跳过空行
    params[key] = value
# ---- 参数区 ----
AAV_name = str(params['AAV_name'])
plasmid_ref = str(params['plasmid_ref'])
AAV_ref = str(params['AAV_ref'])
RepCap_name = str(params['RepCap_name'])
pHelper_name = str(params['pHelper_name'])
hg_name = str(params['hg_name'])

# ---- 文件夹区 ----
HG_FA = glob.glob(os.path.abspath(f"./reference/{hg_name}.*"))[0]
CODE_DIR = os.path.abspath("./code")
OUTPUT = os.path.abspath("./output")
BACKBONE = os.path.join(OUTPUT, "backbone")
HG = os.path.join(OUTPUT, "hg")
HG_NEW = os.path.join(HG, "new")
HG_NEW_OUT = os.path.join(HG_NEW, "output")
IGV = os.path.abspath("./IGV")

def mkdir_if_missing(path):
    if not os.path.exists(path):
        os.makedirs(path)

def copy(src, dst):
    # dst 可以是目录或具体文件路径
    if os.path.isdir(dst):
        shutil.copy(src, dst)
    else:
        shutil.copy(src, os.path.join(dst, os.path.basename(src)))

def run_script(path, args=None, cwd=None):
    cmd = [sys.executable, path]
    if args:
        cmd += args
    print(f"→ Running: {' '.join(cmd)}  (cwd={cwd or os.getcwd()})")
    subprocess.check_call(cmd, cwd=cwd)

def main():
    # 1. 创建目录
    for d in [OUTPUT, BACKBONE, HG, HG_NEW, HG_NEW_OUT, IGV]:
        mkdir_if_missing(d)

    # 2. 复制脚本
    copies = [
        (f"{CODE_DIR}/demul-by-match-score.py", BACKBONE),
        (f"{CODE_DIR}/demul-by-match-score.py", "."),  
        (f"{CODE_DIR}/demul-by-match-score.py", HG_NEW),
        (f"{CODE_DIR}/blat.py", HG),
        (f"{CODE_DIR}/blat.py", HG_NEW_OUT),
        (f"{CODE_DIR}/calculate.py", HG_NEW_OUT),
        (f"{CODE_DIR}/length_static.py", HG_NEW_OUT),
        (f"{CODE_DIR}/filter_from_excel.py", HG_NEW_OUT),
        (f"{CODE_DIR}/filter_from_excel.py", HG),
    ]
    for src, dst in copies:
        copy(src, dst)

    # 3. 移动 AAV_noBB.docx
    shutil.move(f"{AAV_name}.docx", BACKBONE)

    # 4. backbone 流程
    run_script("demul-by-match-score.py")  # 根目录
    shutil.copy(f"{OUTPUT}/backbone.fastq", BACKBONE)
    run_script("demul-by-match-score.py", cwd=BACKBONE)
    bb_out_fastq = os.path.join(BACKBONE, "output", f"{AAV_name}.fastq")
    if not os.path.exists(bb_out_fastq):
        raise FileNotFoundError(f"期望文件不存在：{bb_out_fastq}")

    # 5. hg 流程
    nomatch = os.path.join(OUTPUT, "nomatch.fastq")
    shutil.move(nomatch, HG)

    run_script("blat.py", args=[HG_FA], cwd=HG)
    run_script("filter_from_excel.py", cwd=HG)

    # 复制到 hg/new
    bb_docx = os.path.join(BACKBONE, f"{AAV_name}.docx")
    if not os.path.exists(bb_docx):
        raise FileNotFoundError(f"未在 backbone 中找到 {AAV_name}.docx")
    shutil.copy(bb_docx, HG_NEW)

    align_hg = os.path.join(HG, "align_to_hg.fastq")
    if not os.path.exists(align_hg):
        raise FileNotFoundError(f"未在 hg 中找到 {align_hg}")
    shutil.copy(align_hg, HG_NEW)

    run_script("demul-by-match-score.py", cwd=HG_NEW)

    # hg/new/output 流程
    run_script("blat.py", args=[HG_FA], cwd=HG_NEW_OUT)
    run_script("filter_from_excel.py", cwd=HG_NEW_OUT)
    run_script("calculate.py", cwd=HG_NEW_OUT)
    run_script("length_static.py", cwd=HG_NEW_OUT)

    # 可视化
    os.rename(os.path.join(BACKBONE, "output", f"{AAV_name}.fastq"), os.path.join(BACKBONE, "output", "bb_chimeric.fastq"))
    os.rename(os.path.join(HG_NEW_OUT, f"{AAV_name}.fastq"), os.path.join(BACKBONE, "output", "hg_chimeric.fastq"))
    os.rename(os.path.join(BACKBONE, "output", "nomatch.fastq"), os.path.join(BACKBONE, "output", "backbone_only.fastq"))
    cases = [
        {
            "ref": os.path.abspath(f"./reference/{plasmid_ref}.fasta"),
            "fqs": [
                {"type": "fixed", "pattern": os.path.join(BACKBONE, "output", "bb_chimeric.fastq")},
                {"type": "fixed", "pattern": os.path.join(BACKBONE, "output", "backbone_only.fastq")},
                {"type": "fixed", "pattern": os.path.join(HG_NEW_OUT, "hg_chimeric.fastq")}
            ]
        },
        {
            "ref": os.path.abspath(f"./{RepCap_name}.fasta"),
            "fqs": [{"type": "fixed", "pattern": os.path.join(OUTPUT, f"{RepCap_name}.fastq")}]
        },
        {
            "ref": os.path.abspath(f"./{pHelper_name}.fasta"),
            "fqs": [{"type": "fixed", "pattern": os.path.join(OUTPUT, f"{pHelper_name}.fastq")}]
        }
    ]

    for case in cases:
        ref_path = case["ref"]
        for fq_spec in case["fqs"]:
            # 处理fq路径
            if fq_spec["type"] == "glob":
                matched = glob.glob(fq_spec["pattern"])
                if not matched:
                    print(f"Warning: No files found for {fq_spec['pattern']}")
                    continue
                fq_path = matched[0]  # 取第一个匹配项
            else:
                fq_path = fq_spec["pattern"]
            
            # 构造命令（处理路径中的空格）
            cmd = ["python", "visualize.py", "--fq", fq_path, "--ref", ref_path, "--out", IGV]
            subprocess.run(cmd, check=True)


    print("🎉 全部步骤执行完毕！")

if __name__ == "__main__":
    main()
